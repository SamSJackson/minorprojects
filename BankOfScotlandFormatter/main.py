import calendar, pendulum, os, argparse
from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from Transaction import Transaction
from TransactionWeek import TransactionWeek


wrap_text = Alignment(wrapText=True)

week_beginning_font = Font(
	bold=True,
	size=20,
	)

grey_background = PatternFill(
	start_color="BFBFBF", 
	end_color="BFBFBF", 
	fill_type = "solid")

green_background = PatternFill(
	start_color="1CDE25", 
	end_color="1CDE25", 
	fill_type = "solid")

incoming_money_background = PatternFill(
	start_color="00B050",
	end_color="00B050",
	fill_type = "solid",
	)

red_background = PatternFill(
	start_color="FF4747", 
	end_color="FF4747", 
	fill_type = "solid"
	)

outgoing_money_background = PatternFill(
	start_color="F54545",
	end_color="F54545",
	fill_type = "solid",
	)

black_border = Border(
	left=Side(style='thin'), 
 	right=Side(style='thin'), 
	top=Side(style='thin'), 
	bottom=Side(style='thin'),
	)

def stretch_columns(formatted_sheet):
	dims = {}
	for row in formatted_sheet.rows:
		for cell in row:
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
	
	for col, value in dims.items():
		formatted_sheet.column_dimensions[col].width = value + 3

def parse_filename(from_date, to_date):
	from_date_string = datetime.combine(from_date, datetime.min.time()).strftime("%d_%m_%y")
	to_date_string = datetime.combine(to_date, datetime.min.time()).strftime("%d_%m_%y")
	return from_date_string + "_to_" + to_date_string + '.xlsx'

def same_week(date1, date2):
	return date1.isocalendar()[1] == date2.isocalendar()[1] \
					and date1.year == date2.year

def get_median_week_outgoing(transaction_weeks):
	outgoing_costs = []
	for transaction_week in transaction_weeks:
		outgoing_costs.append(transaction_week.outgoing_total)
	median_index = len(outgoing_costs)//2
	return outgoing_costs[median_index]

def get_mean_week_outgoing(transaction_weeks):
	outgoing_costs_mean = 0
	for transaction_week in transaction_weeks:
		outgoing_costs_mean += transaction_week.outgoing_total
	outgoing_costs_mean /= len(transaction_weeks)
	return outgoing_costs_mean

def get_week_beginning(date):
	datetime_instance = datetime.combine(date, datetime.min.time())
	pendulum_instance = pendulum.instance(datetime_instance)
	pendulum_wk_start = pendulum_instance.start_of('week')
	datetime_string = pendulum_wk_start.to_datetime_string()
	week_start_dt = datetime.fromisoformat(datetime_string)
	return datetime.date(week_start_dt)

def add_week_beginning_to_sheet(sheet, week):
	date = get_week_beginning(week.get_first_transaction_date())
	datetime_instance = datetime.combine(date, datetime.min.time())
	newest_row = sheet.max_row
	value = f"Week Beginning {datetime_instance.strftime('%d/%m/%y')}"
	date_cell = sheet.cell(row=newest_row+1, column=1, value=value)
	date_cell.font = week_beginning_font

def add_median_outgoing_to_sheet(sheet, transactions):
	column = 11
	header_row = 2 
	value_row = header_row + 1
	median_outgoing = get_median_week_outgoing(transactions)
	header_cell = sheet.cell(row=header_row, column=column, value=f"Median Weekly Outgoing")
	value_cell = sheet.cell(row=value_row, column=column, value=median_outgoing)
	value_cell.number_format = '£#,##0.00;-£#,##0.00'
	value_cell.fill = red_background

def add_mean_outgoing_to_sheet(sheet, transactions):
	column = 12
	header_row = 2 
	value_row = header_row + 1
	mean_outgoing = get_mean_week_outgoing(transactions)
	header_cell = sheet.cell(row=header_row, column=column, value=f"Mean Weekly Outgoing")
	value_cell = sheet.cell(row=value_row, column=column, value=mean_outgoing)
	value_cell.number_format = '£#,##0.00;-£#,##0.00'
	value_cell.fill = red_background

def add_transaction_to_sheet(sheet, transaction):
	row = sheet.max_row + 1
	date = str(transaction.date)
	date_day = transaction.date.strftime("%a")
	date_cell = sheet.cell(row=row, column=1, value=f"{date_day} {date}")
	type_cell = sheet.cell(row=row, column=2, value=f"{transaction.transaction_type}")
	info_cell = sheet.cell(row=row, column=3, value=f"{transaction.description}")
	amount_cell = sheet.cell(row=row, column=4, value=transaction.amount)
	amount_cell.fill = outgoing_money_background if transaction.from_account else incoming_money_background
	amount_cell.number_format = '£#,##0.00;-£#,##0.00'
	total_bank_balance_cell = sheet.cell(row=row, column=5, value=transaction.balance)
	total_bank_balance_cell.number_format = '£#,##0.00;-£#,##0.00'

def add_week_sum_to_sheet(sheet, week):
	row = sheet.max_row + 1
	incoming_cell = sheet.cell(row=row, column=6, value=week.income_total)
	incoming_cell.number_format = '£#,##0.00;-£#,##0.00'
	incoming_cell.fill = green_background

	outgoing_cell = sheet.cell(row=row, column=7, value=week.outgoing_total)
	incoming_cell.number_format = '£#,##0.00;-£#,##0.00'
	outgoing_cell.fill = red_background

	sheet.append(["" for x in range(6)])

def parse_line(line) -> list:
	returned_values = []
	split_line = line.split(",")
	transaction_info = {'date': split_line[0], 
				'transaction_type': split_line[1],
				'description': split_line[4],
				'amount_out': split_line[5],
				'amount_in': split_line[6],
				'balance': split_line[7].strip()
				}
	return Transaction(transaction_info['date'], transaction_info['transaction_type'],
						transaction_info['description'], (transaction_info['amount_out'], transaction_info['amount_in']),
						transaction_info['balance'])

def parse_into_weeks(transactions : list) -> list:
	if (len(transactions)) == 0:
		raise ValueError("Transactions empty")

	week_separated_transactions = [TransactionWeek([transactions[0]])]
	last_day = transactions[0].date
	for transaction in transactions[1:]:
		if (same_week(last_day, transaction.date)):
			week_separated_transactions[-1].add_to_week(transaction)
		else:
			week_separated_transactions.append(TransactionWeek([transaction]))
			last_day = transaction.date
	return week_separated_transactions


def read_csv(filename):
	all_transanctions = []
	with open(filename) as file:
		lines = file.readlines()
		for line in lines[1:]:
			transaction = parse_line(line)
			all_transanctions.append(transaction)

	week_separated_transactions = parse_into_weeks(all_transanctions)
	return week_separated_transactions


def write_to_excel(week_separated_transactions : list):
	from_date = week_separated_transactions[-1].get_last_transaction_date()
	to_date = week_separated_transactions[0].get_first_transaction_date()
	formatted_filename = parse_filename(from_date, to_date)
	
	workbook = Workbook()
	sheet = workbook.active

	sheet.append([f"{calendar.month_name[from_date.month]} {from_date.year}", "", "", "", "", "", f"{calendar.month_name[to_date.month]} {to_date.year}"])
	sheet.append(["Date", "Transaction Type", "Description", "Amount", "Balance", "Weekly In", "Weekly Out"])

	add_median_outgoing_to_sheet(sheet, week_separated_transactions)
	add_mean_outgoing_to_sheet(sheet, week_separated_transactions)

	for cell in sheet[2]:
		if cell.value:
			cell.fill = grey_background
			cell.border = black_border

	for transactionWeek in week_separated_transactions:
		add_week_beginning_to_sheet(sheet, transactionWeek)
		for transaction in transactionWeek.transactions:
			add_transaction_to_sheet(sheet, transaction)
		add_week_sum_to_sheet(sheet, transactionWeek)

	stretch_columns(sheet)

	output_path = os.getcwd() + '/Formatted Statements/'
	if not (os.path.exists(output_path)):
		os.makedirs(output_path)
	workbook.save(output_path + formatted_filename)
	

if __name__ == '__main__':
	parser = argparse.ArgumentParser(description="Format transaction statements")
	parser.add_argument('-f', '--filename', metavar='Filename', type=str, nargs='+',
                    help='A filename for formatting csv transaction file')

	args = parser.parse_args()
	filename = vars(args)["filename"] if type(vars(args)["filename"]) == str else vars(args)["filename"][0]
	filename = filename if filename.endswith('csv') else filename + ".csv"

	if not (os.path.exists(filename)):
		raise ValueError("File not found")

	weeks_parsed = read_csv(filename)
	write_to_excel(weeks_parsed)
		

