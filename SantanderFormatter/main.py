import calendar, pendulum
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from transaction import Transaction
from transaction_week import TransactionWeek


wrap_text = Alignment(wrapText=True)

week_beginning_font = Font(
	bold=True,
	size=20,
	)

grey_background = PatternFill(
	start_color="BFBFBF", 
	end_color="BFBFBF", 
	fill_type = "solid")

green_background = PatternFill(
	start_color="1CDE25", 
	end_color="1CDE25", 
	fill_type = "solid")

incoming_money_background = PatternFill(
	start_color="00B050",
	end_color="00B050",
	fill_type = "solid",
	)

red_background = PatternFill(
	start_color="FF4747", 
	end_color="FF4747", 
	fill_type = "solid"
	)

outgoing_money_background = PatternFill(
	start_color="F54545",
	end_color="F54545",
	fill_type = "solid",
	)

black_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'),
               		)


def stretch_columns(formatted_sheet):
	dims = {}
	for row in formatted_sheet.rows:
		for cell in row:
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
	
	for col, value in dims.items():
		formatted_sheet.column_dimensions[col].width = value + 3

def same_week(date1, date2):
	return date1.isocalendar()[1] == date2.isocalendar()[1] \
					and date1.year == date2.year

def get_week_beginning(date):
	datetime_instance = datetime.combine(date, datetime.min.time())
	pendulum_instance = pendulum.instance(datetime_instance)
	pendulum_wk_start = pendulum_instance.start_of('week')
	datetime_string = pendulum_wk_start.to_datetime_string()
	datetime_instance = datetime.fromisoformat(datetime_string)
	return datetime.date(datetime_instance)

def add_week_beginning(sheet, date):
	datetime_instance = datetime.combine(date, datetime.min.time())
	newest_row = sheet.max_row
	value = f"Week Beginning {datetime_instance.strftime('%d/%m/%y')}"
	date_cell = sheet.cell(row=newest_row+1, column=1, value=value)
	date_cell.font = week_beginning_font

def add_transactions_row(sheet, transaction):
	transaction_list = transaction.return_transaction_list()
	newest_row = sheet.max_row
	date_cell = sheet.cell(row=newest_row+1, column=1, value=transaction_list[0])
	info_cell = sheet.cell(row=newest_row+1, column=2, value=transaction_list[1])
	other_party_cell = sheet.cell(row=newest_row+1, column=3, value=transaction_list[2])
	amount_cell = sheet.cell(row=newest_row+1, column=4, value=transaction_list[3])
	amount_cell.fill = outgoing_money_background if transaction.is_from_account else incoming_money_background
	amount_cell.number_format = '£#,##0.00;-£#,##0.00'
	total_bank_balance = sheet.cell(row=newest_row+1, column=5, value=transaction_list[4])

def parse_filename(sheet):
	file_dates = sheet["D2"].value
	file_dates = file_dates.replace("/", "_")
	file_dates = file_dates.replace(" ", "_")
	return file_dates + '.xlsx'

def get_all_transactions(sheet):
	all_transactions = []
	final_row = sheet.max_row - 1 
	transactions = sheet[f"B6:H{final_row}"]
	return transactions

def parse_into_weeks(sheet):
	parsed_weeks = []
	all_transactions = get_all_transactions(sheet)
	if not (len(all_transactions)):
		return

	last_day = all_transactions[0][0].value
	all_transactions[0][0].value = datetime.date(all_transactions[0][0].value)
	parsed_weeks.append([list(all_transactions[0])])
	for transaction in all_transactions[1:]:
		date = transaction[0].value
		transaction[0].value = datetime.date(transaction[0].value)
		if (same_week(last_day, date)):
			parsed_weeks[-1].append(list(transaction))
		else:
			parsed_weeks.append([list(transaction)])
			last_day = date
	return parsed_weeks

def clean_weeks(sheet):
	parsed_weeks = parse_into_weeks(sheet)
	for week in parsed_weeks:
		for index, transaction in enumerate(week):
			date = transaction[0].value
			info = transaction[2].value
			amount = transaction[4].value if transaction[4].value else (transaction[5].value * -1)
			total_balance = transaction[6].value
			week[index] = Transaction(date, info, amount, total_balance)
		week = [transaction for transaction in week if transaction.info == 'CARD']
	return parsed_weeks

def sum_weeks(cleaned_weeks : list):
	weeks_costs = []
	for week in cleaned_weeks:
		transactionWeek = TransactionWeek(week)
		cost_tuple = transactionWeek.month_total()
		weeks_costs.append(cost_tuple)
	return weeks_costs

def write_to_sheet(sheet, formatted_sheet, cleaned_weeks, weeks_sum):
	file_dates = sheet["D2"].value.split(" ")
	from_date, to_date = datetime.date(datetime.strptime(file_dates[0], "%d/%m/%Y")), datetime.date(datetime.strptime(file_dates[2], "%d/%m/%Y"))
	formatted_sheet.append([f"{calendar.month_name[from_date.month]} {from_date.year}", "", "", "",f"{calendar.month_name[to_date.month]} {to_date.year}"])
	formatted_sheet.append([f"Date", "Info", "Other Party", "Amount","Total Bank Balance", "Weekly In", "Weekly Out", ""])
	
	for cell in formatted_sheet[2]:
		cell.fill = grey_background
		cell.border = black_border

	for index, week in enumerate(cleaned_weeks):
		date_beginning = get_week_beginning(week[-1].date)
		add_week_beginning(formatted_sheet, date_beginning)
		for transaction in week:
			add_transactions_row(formatted_sheet, transaction)
		formatted_sheet.append(["", "", "", "", "", weeks_sum[index][0], weeks_sum[index][1]])
		formatted_sheet.append([])

	for row in formatted_sheet:
		if row == formatted_sheet[1] or row == formatted_sheet[2]:
			continue
		f_cell = row[-3]
		g_cell = row[-2]
		if f_cell.value or f_cell.value == 0:
			f_cell.fill = green_background
		
		if g_cell.value or g_cell.value == 0:
			g_cell.fill = red_background

	stretch_columns(formatted_sheet)


if __name__ == '__main__':
	workbook = load_workbook(filename='Statements09012960416133.xlsx')
	sheet = workbook.active

	formatted_workbook = Workbook()
	formatted_sheet = formatted_workbook.active
	formatted_workbook_name = parse_filename(sheet)

	cleaned_weeks = clean_weeks(sheet)
	weeks_sum = sum_weeks(cleaned_weeks)
	write_to_sheet(sheet, formatted_sheet, cleaned_weeks, weeks_sum)

	formatted_workbook.save(filename=formatted_workbook_name)

